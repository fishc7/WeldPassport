"""Импорт сварщиков из табелей Excel в РАБОТНИКИ + СВАРЩИКИ.

Сканирует папки из config_hr.json, определяет формат каждого файла,
извлекает сварщиков и записывает новых в БД.

Поддерживаемые форматы:
  B — ОГС / офис-отчёт: заголовок «Ф.И.О.», нет колонки Должность
      → все строки считаются сварщиками
  C — Табель с заголовками «ФИО» и «Должность» в разных колонках
  D — Компактный список без заголовков: пусто | ФИО | Должность
  A — T-13 (Унифицированная форма): пропускается, должность не извлекается

Использование:
    python scripts/import_welders.py --report
        Сканирует файлы, генерирует Excel для проверки. БД не трогает.

    python scripts/import_welders.py --from-report review_2026-06-21.xlsx
        Читает ваши решения из Excel и импортирует в БД.

    python scripts/import_welders.py --dry-run
        Сканирует и показывает кандидатов без записи в БД.

    python scripts/import_welders.py --rescan
        Игнорирует лог, пересканирует все файлы заново.

    python scripts/import_welders.py path/to.xlsx
        Обработать конкретный файл.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd

from config import settings
from db import get_connection

# ── Пути ────────────────────────────────────────────────────────────────────

CONFIG_PATH = Path(__file__).resolve().parents[1] / "config_hr.json"
LOG_PATH = Path(r"D:\WeldPassport\Работники\import_log.json")


# ── Конфиг и лог ────────────────────────────────────────────────────────────

def load_config() -> dict:
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    default = {
        "scan_dirs": [r"D:\WeldPassport\Работники\Табеля"],
        "organizatsiya": "МК Кран",
        "welder_titles": ["эл/сварщик", "эл/сварщик тт", "сварщик", "электросварщик"],
    }
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(default, f, ensure_ascii=False, indent=2)
    print(f"Создан конфиг по умолчанию: {CONFIG_PATH}")
    return default


def load_log() -> dict:
    if LOG_PATH.exists():
        with open(LOG_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_log(log: dict) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def _file_key(path: Path) -> str:
    return str(path).lower()


def is_processed(path: Path, log: dict) -> bool:
    entry = log.get(_file_key(path))
    if not entry:
        return False
    return entry.get("mtime") == path.stat().st_mtime


# ── Сканирование ─────────────────────────────────────────────────────────────

def scan_files(dirs: list[str], name_filter: str = "") -> list[Path]:
    found: list[Path] = []
    seen: set[str] = set()
    for d in dirs:
        p = Path(d)
        if not p.exists():
            print(f"  ПАПКА НЕ НАЙДЕНА: {d}")
            continue
        for ext in ("*.xlsx", "*.xls"):
            for f in p.rglob(ext):
                if f.name.startswith("~"):
                    continue
                if name_filter and name_filter.lower() not in f.name.lower():
                    continue
                key = _file_key(f)
                if key not in seen:
                    seen.add(key)
                    found.append(f)
    return sorted(found)


# ── Детектор и парсер форматов ───────────────────────────────────────────────

def _cell(v) -> str:
    if not pd.notna(v):
        return ""
    return str(v).strip()


def _low(s: str) -> str:
    return s.lower()


def _contains(cell: str, *keywords: str) -> bool:
    low = _low(cell)
    return any(k in low for k in keywords)


def _looks_like_name(v: str) -> bool:
    v = normalize_fio(v)
    parts = v.split()
    return (
        len(parts) >= 2
        and bool(re.search(r"[А-ЯЁа-яё]", v))
        and not any(c.isdigit() for c in v)
        and len(v) > 5
    )


class SheetResult:
    def __init__(self, fmt: str, rows: list[tuple[str, str]], sheet: str):
        self.fmt = fmt
        self.rows = rows
        self.sheet = sheet


def _detect_sheet(df: pd.DataFrame, sheet: str) -> SheetResult:
    """Определить формат листа и вернуть найденные пары (ФИО, Должность)."""

    for i, row in df.iterrows():
        row_cells = {j: _cell(v) for j, v in enumerate(row)}
        nonempty = {j: v for j, v in row_cells.items() if v}

        # T-13: "Фамилия, \nинициалы, \nдолжность" — в ОДНОЙ ячейке
        for v in nonempty.values():
            if _contains(v, "фамилия") and _contains(v, "должность", "специальность", "профессия"):
                return SheetResult("A", [], sheet)

        # Format C: «ФИО» и «Должность» в РАЗНЫХ ячейках одной строки
        fio_col = next(
            (j for j, v in nonempty.items()
             if _contains(v, "ф.и.о", "фио", "фамилия_имя", "фамилия")),
            None,
        )
        dol_col = next(
            (j for j, v in nonempty.items()
             if _contains(v, "должность", "специальность")),
            None,
        )
        if fio_col is not None and dol_col is not None and fio_col != dol_col:
            rows = _parse_c(df, fio_col, dol_col, data_start=i)
            if rows:
                return SheetResult("C", rows, sheet)

        # Format B: есть «Ф.И.О.» но нет отдельной «Должность»
        if fio_col is not None and dol_col is None:
            rows = _parse_b(df, fio_col, data_start=i)
            if rows:
                return SheetResult("B", rows, sheet)

    # Format D: нет заголовков — компактный список
    rows = _parse_d(df)
    if rows:
        return SheetResult("D", rows, sheet)

    return SheetResult("?", [], sheet)


def _parse_c(df: pd.DataFrame, fio_col: int, dol_col: int, data_start: int) -> list[tuple[str, str]]:
    results = []
    for i, row in df.iterrows():
        if i <= data_start:
            continue
        cells = [_cell(v) for v in row]
        fio = cells[fio_col] if fio_col < len(cells) else ""
        dol = cells[dol_col] if dol_col < len(cells) else ""
        if _looks_like_name(fio) and dol:
            results.append((fio, dol))
    return results


def _parse_b(df: pd.DataFrame, fio_col: int, data_start: int) -> list[tuple[str, str]]:
    results = []
    for i, row in df.iterrows():
        if i <= data_start:
            continue
        cells = [_cell(v) for v in row]
        # Строки данных: первая колонка — порядковый номер
        first = cells[0] if cells else ""
        fio = cells[fio_col] if fio_col < len(cells) else ""
        if first.isdigit() and _looks_like_name(fio):
            results.append((fio, "эл/сварщик"))
    return results


def _parse_d(df: pd.DataFrame) -> list[tuple[str, str]]:
    results = []
    for _, row in df.iterrows():
        cells = [_cell(v) for v in row]
        if len(cells) >= 3 and not cells[0] and _looks_like_name(cells[1]) and cells[2]:
            results.append((cells[1], cells[2]))
    return results


def parse_file(path: Path) -> list[SheetResult]:
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        print(f"    ОШИБКА чтения файла: {e}")
        return []

    results = []
    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=str)
        except Exception as e:
            print(f"    ОШИБКА лист «{sheet}»: {e}")
            continue
        results.append(_detect_sheet(df, sheet))
    return results


# ── Фильтр сварщиков ─────────────────────────────────────────────────────────

def filter_welders(rows: list[tuple[str, str]], w_set: set[str]) -> list[tuple[str, str]]:
    return [(fio, dol) for fio, dol in rows if dol.strip().lower() in w_set]


# ── Нормализация и дедупликация ───────────────────────────────────────────────

def normalize_fio(fio: str) -> str:
    """Нормализует ФИО: лишние пробелы, переносы, ё→е, точки в конце."""
    s = fio.strip()
    s = re.sub(r'[\r\n\t]+', ' ', s)       # переносы строк → пробел
    s = re.sub(r' {2,}', ' ', s)            # множественные пробелы
    s = s.rstrip('.,;')                      # точки/запятые в конце
    s = s.replace('ё', 'е').replace('Ё', 'Е')
    return s.strip()


def _surname(fio: str) -> str:
    parts = normalize_fio(fio).split()
    return parts[0].lower() if parts else ""


def _first_initial(fio: str) -> str:
    """Первая буква имени (второе слово)."""
    parts = normalize_fio(fio).split()
    return parts[1][0].lower() if len(parts) >= 2 else ""


def _find_similar(fio: str, existing: set[str]) -> list[str]:
    """
    Совпадение по фамилии + первая буква имени.
    Ловит: 'Алашеев С.С.' ↔ 'Алашеев Сергей Сергеевич'
           'Иванов А.В.'  ↔ 'Иванов Анатолий Владимирович'
    """
    fio_norm = normalize_fio(fio).lower()
    sur  = _surname(fio)
    init = _first_initial(fio)

    matches = []
    for e in existing:
        if normalize_fio(e).lower() == fio_norm:
            continue                          # точное совпадение — не "похожее"
        if _surname(e) == sur and (not init or not _first_initial(e) or _first_initial(e) == init):
            matches.append(e)
    return matches


def is_exact_dupe(fio: str, existing_fio: set[str]) -> bool:
    """Точное совпадение после нормализации."""
    fio_norm = normalize_fio(fio).lower()
    return any(normalize_fio(e).lower() == fio_norm for e in existing_fio)


# ── БД ───────────────────────────────────────────────────────────────────────

def get_existing(conn) -> dict[str, int]:
    with conn.cursor() as cur:
        cur.execute(f'SELECT "ID_Работника","ФИО" FROM "{settings.db_schema}"."РАБОТНИКИ"')
        return {r["ФИО"]: r["ID_Работника"] for r in cur.fetchall()}


def insert_rabotnik(conn, fio: str, dol: str, org: str) -> int:
    with conn.cursor() as cur:
        cur.execute(
            f"""INSERT INTO "{settings.db_schema}"."РАБОТНИКИ"
                ("ФИО","Должность","Организация","Статус")
                VALUES (%s,%s,%s,'активный') RETURNING "ID_Работника" """,
            (fio, dol, org),
        )
        return cur.fetchone()["ID_Работника"]


def insert_svarshchik(conn, id_rab: int) -> None:
    with conn.cursor() as cur:
        cur.execute(
            f"""INSERT INTO "{settings.db_schema}"."СВАРЩИКИ"
                ("ID_Работника","Статус_Сварщика") VALUES (%s,'активный')""",
            (id_rab,),
        )


# ── Main ─────────────────────────────────────────────────────────────────────

FMT_LABEL = {
    "A": "T-13 (пропущен)",
    "B": "ОГС/офис (все — сварщики)",
    "C": "Табель ФИО+Должность",
    "D": "Компактный список",
    "?": "формат не распознан",
}

REPORT_DIR = Path(r"D:\WeldPassport\Работники")

# Цвета для Excel
_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # новый
_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # похожее имя
_HEADER = PatternFill("solid", fgColor="1F4E79")   # заголовок
_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

REPORT_COLS = [
    ("Действие",          20),   # A — пользователь заполняет
    ("Статус",            18),   # B
    ("ФИО",               35),   # C — из файла
    ("ФИО_исправленное",  35),   # D — пользователь заполняет если опечатка
    ("Должность",         20),   # E
    ("Источник",          35),   # F
    ("Похожее_в_БД",      45),   # G
]


# ── Генерация Excel-отчёта ───────────────────────────────────────────────────

def generate_report(
    to_add: dict[str, str],
    similar_map: dict[str, list[str]],
    source_map: dict[str, str],   # fio → имя файла-источника
) -> Path:
    today = datetime.now().strftime("%Y-%m-%d")
    path  = REPORT_DIR / f"review_{today}.xlsx"
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "К проверке"

    # ── Заголовок ──
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (title, width) in enumerate(REPORT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill   = _HEADER
        cell.font   = header_font
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Данные ──
    for row_idx, (fio, dol) in enumerate(sorted(to_add.items()), start=2):
        is_similar = fio in similar_map
        fill       = _YELLOW if is_similar else _GREEN
        status     = "Похожее имя" if is_similar else "Новый"
        действие   = "уточнить"   if is_similar else "добавить"
        похожие    = " | ".join(similar_map.get(fio, []))
        источник   = source_map.get(fio, "")

        values = [действие, status, fio, "", dol, источник, похожие]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = _THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Колонка A — выпадающий список для удобства
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"добавить,пропустить,уточнить"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"A2:A{ws.max_row}"

    # Пояснение на отдельном листе
    ws_info = wb.create_sheet("Инструкция")
    info = [
        ("Колонка",           "Описание"),
        ("Действие",          "добавить — импортировать в БД"),
        ("",                  "пропустить — не импортировать"),
        ("",                  "уточнить — требует проверки (похожее имя уже есть)"),
        ("ФИО_исправленное",  "Заполните если в ФИО опечатка. Импорт возьмёт исправленное имя."),
        ("Статус",            "Новый — в БД нет похожих. Похожее имя — есть похожий человек в БД."),
        ("Похожее_в_БД",      "Кто уже есть в базе с похожим именем."),
    ]
    for r, (a, b) in enumerate(info, start=1):
        ws_info.cell(r, 1, a).font = Font(bold=(r == 1))
        ws_info.cell(r, 2, b)
    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 60

    wb.save(path)
    return path


# ── Импорт из Excel-отчёта ───────────────────────────────────────────────────

def import_from_report(report_path: Path, org: str, dry_run: bool, auto_yes: bool = False) -> int:
    print(f"Читаю отчёт: {report_path.name}")

    wb = openpyxl.load_workbook(report_path)
    ws = wb["К проверке"]

    # Считываем заголовки
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    to_import: list[tuple[str, str]] = []

    for row in range(2, ws.max_row + 1):
        def v(colname):
            c = col.get(colname)
            val = ws.cell(row, c).value if c else None
            return str(val).strip() if val else ""

        действие = v("Действие").lower()
        if действие != "добавить":
            continue

        fio = v("ФИО_исправленное") or v("ФИО")
        dol = v("Должность")
        if fio and dol:
            to_import.append((fio, dol))

    print(f"Строк к импорту: {len(to_import)}")
    if not to_import:
        print("Нет строк с Действие='добавить'.")
        return 0

    # Сравнение с БД
    with get_connection() as conn:
        existing = get_existing(conn)
    existing_fio = set(existing.keys())

    new_rows  = [(fio, dol) for fio, dol in to_import if fio not in existing_fio]
    dupe_rows = [(fio, dol) for fio, dol in to_import if fio in existing_fio]

    if dupe_rows:
        print(f"Уже в БД ({len(dupe_rows)}) — пропущены:")
        for fio, _ in dupe_rows:
            print(f"  = {fio}")

    print(f"\nБудет добавлено: {len(new_rows)}")
    for fio, dol in new_rows:
        print(f"  + {fio} ({dol})")

    if dry_run:
        print("\n[--dry-run] Запись в БД не выполнялась.")
        return 0

    if not new_rows:
        print("Нечего добавлять.")
        return 0

    if auto_yes:
        print("\nПродолжить импорт? [y/N]: y (--yes)")
    else:
        confirm = input("\nПродолжить импорт? [y/N]: ").strip().lower()
        if confirm not in ("y", "д"):
            print("Отменено.")
            return 0

    inserted = 0
    with get_connection() as conn:
        for fio, dol in new_rows:
            insert_rabotnik(conn, fio, dol, org)
            id_rab = get_existing(conn)[fio]
            insert_svarshchik(conn, id_rab)
            print(f"  ДОБАВЛЕН: {fio}")
            inserted += 1

    print(f"\nГотово. Добавлено: {inserted}.")
    return 0


def main() -> int:
    dry_run     = "--dry-run"     in sys.argv
    rescan      = "--rescan"      in sys.argv
    make_report = "--report"      in sys.argv
    auto_yes    = "--yes"         in sys.argv
    from_report = next((sys.argv[i + 1] for i, a in enumerate(sys.argv)
                        if a == "--from-report" and i + 1 < len(sys.argv)), None)

    file_args = [Path(a) for a in sys.argv[1:]
                 if not a.startswith("--") and a != from_report]

    cfg   = load_config()
    log   = load_log()
    org   = cfg.get("organizatsiya", "")
    w_set = {t.lower() for t in cfg.get("welder_titles", [])}

    # ── Режим: импорт из готового отчёта ──
    if from_report:
        return import_from_report(Path(from_report), org, dry_run, auto_yes)

    # ── Файлы для обработки ──
    if file_args:
        files = file_args
        skip_count = 0
    else:
        all_files  = scan_files(cfg.get("scan_dirs", []), cfg.get("file_name_contains", ""))
        files      = [f for f in all_files if not is_processed(f, log) or rescan]
        skip_count = len(all_files) - len(files)

    print(f"\nФайлов к обработке: {len(files)}"
          + (f"  (пропущено в логе: {skip_count})" if skip_count else ""))

    if not files:
        print("Новых файлов нет. Используйте --rescan чтобы пересканировать все.")
        return 0

    # ── Сбор кандидатов ──
    # fio → (dolzhnost, source_filename)
    candidates: dict[str, tuple[str, str]] = {}
    file_weld_count: dict[str, int] = {}

    for f in files:
        print(f"\n  {f.name}")
        sheet_results = parse_file(f)
        file_total = 0

        for sr in sheet_results:
            label = FMT_LABEL.get(sr.fmt, sr.fmt)
            if sr.fmt == "A":
                print(f"    Лист «{sr.sheet}»: {label}")
                continue
            if sr.fmt == "?":
                if sr.sheet:
                    print(f"    Лист «{sr.sheet}»: {label}")
                continue

            welders = filter_welders(sr.rows, w_set)
            print(f"    Лист «{sr.sheet}» [{label}]: "
                  f"строк={len(sr.rows)}, сварщиков={len(welders)}")

            for fio, dol in welders:
                fio = normalize_fio(fio)
                if fio not in candidates:
                    candidates[fio] = (dol, f.name)
            file_total += len(welders)

        file_weld_count[_file_key(f)] = file_total

    # ── Отчёт по собранным данным ──
    unique: dict[str, str] = {fio: dol for fio, (dol, _) in candidates.items()}
    print(f"\nУникальных сварщиков во всех файлах: {len(unique)}")

    if not unique:
        print("Нечего импортировать.")
        _mark_processed(log, files, file_weld_count, dry_run)
        return 0

    # ── Сравнение с БД ──
    with get_connection() as conn:
        existing = get_existing(conn)
    existing_fio = set(existing.keys())

    exact_dupes  = {fio for fio in unique if is_exact_dupe(fio, existing_fio)}
    to_add: dict[str, str] = {}
    similar_map:  dict[str, list[str]] = {}

    for fio, dol in unique.items():
        if fio in exact_dupes:
            continue
        sim = _find_similar(fio, existing_fio)
        if sim:
            similar_map[fio] = sim
        to_add[fio] = dol

    # ── Вывод ──
    if exact_dupes:
        print(f"\nУже в БД ({len(exact_dupes)}) — пропущены:")
        for fio in sorted(exact_dupes):
            print(f"  = {fio}")

    if similar_map:
        print(f"\nПохожие имена ({len(similar_map)}) — возможные опечатки:")
        for fio, hits in similar_map.items():
            print(f"  ? {fio!r}  <->  {hits}")

    print(f"\nКандидаты на добавление: {len(to_add)}")
    for fio, dol in sorted(to_add.items()):
        marker = "[!]" if fio in similar_map else "+"
        print(f"  {marker} {fio} ({dol})")

    if dry_run and not make_report:
        print("\n[--dry-run] Запись в БД не выполнялась.")
        return 0

    # ── Режим: сгенерировать Excel-отчёт ──
    if make_report:
        source_map = {fio: src for fio, (_, src) in candidates.items()}
        report_path = generate_report(to_add, similar_map, source_map)
        print(f"\nОтчёт сохранён: {report_path}")
        print("Заполните колонку 'Действие' и при необходимости 'ФИО_исправленное'.")
        print(f"Затем запустите:")
        print(f"  python scripts/import_welders.py --from-report \"{report_path}\"")
        return 0

    if not to_add:
        print("Новых записей нет.")
        _mark_processed(log, files, file_weld_count, dry_run)
        return 0

    confirm = input("\nПродолжить импорт? [y/N]: ").strip().lower()
    if confirm not in ("y", "д"):
        print("Отменено.")
        return 0

    # ── Запись в БД ──
    inserted = skipped_by_user = 0

    with get_connection() as conn:
        for fio, dol in to_add.items():
            if fio in similar_map:
                print(f"\n  [!] Похожее имя уже в БД:")
                print(f"     Новый:  {fio}")
                for s in similar_map[fio]:
                    print(f"     В базе: {s}")
                ans = input("     Добавить? [y/N]: ").strip().lower()
                if ans not in ("y", "д"):
                    skipped_by_user += 1
                    continue
            id_rab = insert_rabotnik(conn, fio, dol, org)
            insert_svarshchik(conn, id_rab)
            inserted += 1

    _mark_processed(log, files, file_weld_count, dry_run)
    print(f"\nГотово. Добавлено: {inserted}, пропущено вами: {skipped_by_user}.")
    return 0


def _mark_processed(log: dict, files: list[Path], counts: dict, dry_run: bool) -> None:
    if dry_run:
        return
    for f in files:
        log[_file_key(f)] = {
            "mtime": f.stat().st_mtime,
            "welders_found": counts.get(_file_key(f), 0),
            "date": datetime.now().isoformat()[:10],
        }
    save_log(log)


if __name__ == "__main__":
    raise SystemExit(main())
